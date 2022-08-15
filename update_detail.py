import openpyxl as x
from mail import sendmail
import math


def update_bmi(name ) :
        work=x.load_workbook('gymdetails.xlsx')
        sheet=work['Sheet1']

        for row in range(1,sheet.max_row+1) :
            if (sheet.cell(row,1)).value == name :
               break

        if sheet.cell(row,1).value != name :
             print("------CANDIDATE NOT EXISTS-----")
             return

        height=float(input("Enter your new height : "))
        weight=float(input("Enter your new weight : "))
        temp = weight / (math.pow(height, 2))
        print(f"Your new BMI is : {temp} ")
        (sheet.cell(row,5)).value=temp
        print("------Details updated--------")

        sendmail(sheet.cell(row,4).value,f"""Hello {sheet.cell(row,1).value.upper()}!
Your new BMI is {temp}""")
        work.save("gymdetails.xlsx")