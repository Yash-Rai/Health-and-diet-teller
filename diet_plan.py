import openpyxl as x

def diet() :
    name = input("Enter your name : ")
    work = x.load_workbook("gymdetails.xlsx")
    sheet = work["Sheet1"]
    for row in range(1, sheet.max_row + 1):
        if (sheet.cell(row, 1)).value == name:
            break

    if sheet.cell(row, 1).value != name:
        print("----YOU ARE NOT REGISTERED YET-----")
        return

    fat=sheet.cell(row,6).value
    if fat=="obese" :
        file=open("obesity.txt",'r')
        plan =file.read()
        print("----YOUR DIET PLAN--------")
        print(plan)

        file.close()

    elif fat=="overweight" :
        file = open("obesity.txt", "r")
        plan = file.read()
        print("----YOUR DIET PLAN--------")
        print(plan)

        file.close()

    elif fat=="fit" :
        file = open("healthy.txt", "r")
        plan = file.read()
        print("----YOUR DIET PLAN--------")
        print(plan)

        file.close()

    elif fat=="underweight" :
        file = open("weightgain.txt", "r")
        plan = file.read()
        print("----YOUR DIET PLAN--------")
        print(plan)

        file.close()

