import openpyxl as x

def print_details() :
        name=input("Enter your name : ")
        work=x.load_workbook("gymdetails.xlsx")
        sheet=work["Sheet1"]
        for row in range(1, sheet.max_row + 1):
            if (sheet.cell(row, 1)).value == name:
                break

        if sheet.cell(row,1).value != name :
                     print("----CANDIDATE NOT EXISTS-----")
                     return


        print(f""" 
        NAME : {sheet.cell(row,1).value.upper()}
        AGE : {sheet.cell(row,2).value}
        GENDER : {sheet.cell(row,3).value.upper()}
        MAIL : {sheet.cell(row,4).value}
        BMI : {sheet.cell(row,5).value}
        FAT : {sheet.cell(row,6).value.upper()}
        DATE OF JOINING : {sheet.cell(row,7).value}
        """)